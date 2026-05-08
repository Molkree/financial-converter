import csv
from abc import ABC, abstractmethod
from collections import defaultdict
from datetime import date
from decimal import Decimal
from itertools import batched
from typing import Any, Literal, Self

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, field_validator, model_validator
from pydantic_extra_types.currency_code import ISO4217

TRANSACTIONS_SHEET = Literal["Expenses", "Income"]
TRANSFERS_SHEET = "Transfers"
BATCH_SIZE = 400


class BaseRow(BaseModel, ABC):
    transaction_date: date
    description: str | None
    to_account: str
    to_amount: Decimal
    to_currency_code: ISO4217

    @field_validator("description", mode="after")
    @classmethod
    def convert_newlines_for_markdown(cls, v: str | None) -> str | None:
        if not v:
            return v
        return v.replace("\n", "  ").strip()

    @classmethod
    @abstractmethod
    def from_excel_row(cls, row: tuple[Any, ...]) -> Self: ...

    @abstractmethod
    def to_tuple(self) -> tuple[Any, ...]: ...


class TransactionRow(BaseRow):
    category: str

    @classmethod
    def from_excel_row(cls, row: tuple[Any, ...]) -> Self:
        return cls(
            transaction_date=row[0],
            category=row[1],
            to_account=row[2],
            to_amount=row[5],
            to_currency_code=row[6],
            description=row[10],
        )

    def to_tuple(self) -> tuple[Any, ...]:
        return (
            self.transaction_date,
            self.description,
            self.to_account,
            self.to_amount,
            self.to_currency_code,
            self.category,
        )

    def __hash__(self) -> int:
        return hash(self.to_tuple())


class TransferRow(BaseRow):
    from_account: str
    from_amount: Decimal
    from_currency_code: ISO4217

    @model_validator(mode="before")
    @classmethod
    def validate_missing_to_amount(cls, data: dict[str, Any]) -> dict[str, Any]:
        if not data.get("to_amount"):
            data["to_amount"] = data.get("from_amount")
            data["to_currency_code"] = data.get("from_currency_code")

        return data

    @classmethod
    def from_excel_row(cls, row: tuple[Any, ...]) -> Self:
        return cls(
            transaction_date=row[0],
            description=row[7],
            from_account=row[1],
            to_account=row[2],
            from_amount=row[3],
            from_currency_code=row[4],
            to_amount=row[5],
            to_currency_code=row[6],
        )

    def to_tuple(self) -> tuple[Any, ...]:
        return (
            self.transaction_date,
            self.description,
            self.from_account,
            self.to_account,
            self.from_amount,
            self.to_amount,
            self.from_currency_code,
            self.to_currency_code,
        )

    def __hash__(self) -> int:
        return hash(self.to_tuple())


def convert_excel_to_csvs(excel_file: str) -> None:
    """Converts the given Excel file into three CSV files: expenses.csv, income.csv and
    transfers.csv."""
    workbook = load_workbook(filename=excel_file)

    for sheet in ("Expenses", "Income"):
        process_transactions_sheet(workbook, sheet)

    process_transfers_sheet(workbook)


def process_transactions_sheet(workbook: Workbook, sheet: TRANSACTIONS_SHEET) -> None:
    worksheet = workbook[sheet]
    headers = (
        "Date",
        "Description",
        "Asset account",
        "Amount",
        "Currency code (ISO 4217)",
        "Category",
    )
    process_sheet(worksheet, sheet, TransactionRow, headers)


def process_transfers_sheet(workbook: Workbook) -> None:
    worksheet = workbook[TRANSFERS_SHEET]
    headers = (
        "Date",
        "Description",
        "From account",
        "To account",
        "From amount",
        "To amount",
        "From currency code (ISO 4217)",
        "To currency code (ISO 4217)",
    )
    process_sheet(worksheet, TRANSFERS_SHEET, TransferRow, headers)


def process_sheet(
    worksheet: Worksheet,
    sheet_name: str,
    row_class: type[BaseRow],
    headers: tuple[str, ...],
) -> None:
    """Generic sheet processor that handles grouping, deduplication, and CSV export."""
    rows_by_date: defaultdict[date, list[BaseRow]] = defaultdict(list)
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        parsed_row = row_class.from_excel_row(row)
        rows_by_date[parsed_row.transaction_date].append(parsed_row)
    converted_rows = [row.to_tuple() for row in uniquify_rows(rows_by_date)]

    csv_index = 1
    for batch in batched(converted_rows, BATCH_SIZE):
        csv_name = f"{sheet_name.lower()}_{csv_index}.csv"
        csv_index += 1

        with open(csv_name, mode="w", encoding="utf-8", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(headers)
            writer.writerows(batch)


def uniquify_rows[T: BaseRow](rows: dict[date, list[T]]) -> list[T]:
    """Enumerates descriptions for identical transactions on the same day."""
    hash_groups: defaultdict[date, defaultdict[int, list[T]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for transaction_date, transactions in rows.items():
        # Group transactions by their hash value to identify duplicates
        for transaction in transactions:
            hash_groups[transaction_date][hash(transaction)].append(transaction)

    for day_groups in hash_groups.values():
        for hash_group in day_groups.values():
            if len(hash_group) > 1:
                for i, transaction in enumerate(hash_group[::-1], start=1):
                    if not transaction.description:
                        transaction.description = f"Transaction {i}"
                    else:
                        transaction.description += f"  Transaction {i}"

    return [
        transaction for transactions in rows.values() for transaction in transactions
    ]


if __name__ == "__main__":
    convert_excel_to_csvs("export.xlsx")
